import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


ALLOWED_SCENE_TAGS = {
    "day", "night", "rain", "intersection", "occlusion",
    "highway", "urban", "turning_scene", "straight_road", "crowded",
}
SCENE_PROFILES = {
    "day_intersection": ["day", "intersection"],
    "day_straight_road": ["day", "straight_road"],
    "day_highway": ["day", "highway"],
    "night_intersection": ["night", "intersection"],
    "night_straight_road": ["night", "straight_road"],
    "rain_intersection": ["rain", "intersection"],
    "rain_straight_road": ["rain", "straight_road"],
    "day_intersection_turning": ["day", "intersection", "turning_scene"],
    "night_intersection_turning": ["night", "intersection", "turning_scene"],
    "custom": [],
}


def parse_scene_tags(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, list):
        tokens = [str(x).strip() for x in v]
    else:
        text = str(v).strip()
        if not text:
            return []
        for sep in ["，", ";", "；", "|", "/", "\\"]:
            text = text.replace(sep, ",")
        tokens = [t.strip() for t in text.split(",") if t.strip()]
    out: List[str] = []
    seen = set()
    for t in tokens:
        if t in ALLOWED_SCENE_TAGS and t not in seen:
            seen.add(t)
            out.append(t)
    return out


def merge_scene_tags(profile: Any, extra: Any, fallback_scene_tags: Any = None) -> List[str]:
    p = str(profile or "").strip()
    tags: List[str] = []
    if p in SCENE_PROFILES:
        tags.extend(SCENE_PROFILES[p])
    extra_tags = parse_scene_tags(extra)
    tags.extend(extra_tags)
    # backward compatibility for old sheet that only has scene_tags
    if not tags:
        tags.extend(parse_scene_tags(fallback_scene_tags))
    dedup: List[str] = []
    seen = set()
    for t in tags:
        if t in ALLOWED_SCENE_TAGS and t not in seen:
            seen.add(t)
            dedup.append(t)
    return dedup


def norm_time(v: Any):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert annotation Excel back to jsonl.")
    parser.add_argument("--input", required=True, help="Input xlsx")
    parser.add_argument("--output", required=True, help="Output jsonl")
    parser.add_argument("--sheet", default="annotation", help="Sheet name")
    args = parser.parse_args()

    inp = Path(args.input).resolve()
    out = Path(args.output).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(inp, data_only=True)
    ws = wb[args.sheet]
    headers = [c.value for c in ws[1]]

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in r):
            continue
        obj = dict(zip(headers, r))
        tags = merge_scene_tags(
            profile=obj.get("scene_profile"),
            extra=obj.get("extra_scene_tags"),
            fallback_scene_tags=obj.get("scene_tags"),
        )

        row = {
            "sample_id": str(obj.get("sample_id", "")).strip(),
            "split": str(obj.get("split", "")).strip(),
            "video": str(obj.get("video", "")).strip(),
            "accident_type": str(obj.get("accident_type", "") or "").strip(),
            "onset_time": norm_time(obj.get("onset_time")),
            "impact_time": norm_time(obj.get("impact_time")),
            "post_time": norm_time(obj.get("post_time")),
            "scene_tags": tags,
            "notes": str(obj.get("notes", "") or "").strip(),
        }
        rows.append(row)

    with out.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    print(f"[DONE] rows={len(rows)} out={out}")


if __name__ == "__main__":
    main()
