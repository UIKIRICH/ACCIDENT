import argparse
import json
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ACCIDENT_TYPES = ["rear_end", "lane_change", "turn_conflict", "generic"]
SCENE_TAGS = [
    "day", "night", "rain", "intersection", "occlusion",
    "highway", "urban", "turning_scene", "straight_road", "crowded",
]
SCENE_PROFILES = {
    "day_intersection": ["day", "intersection"],
    "day_straight_road": ["day", "straight_road"],
    "day_highway": ["day", "highway"],
    "night_intersection": ["night", "intersection"],
    "night_straight_road": ["night", "straight_road"],
    "rain_intersection": ["rain", "intersection"],
    "rain_straight_road": ["rain", "straight_road"],
    "day_intersection_turning": ["day", "intersection", "turning_scene"],
    "night_intersection_turning": ["night", "intersection", "turning_scene"],
    "custom": [],
}
HEADERS = [
    "sample_id", "split", "video", "accident_type",
    "onset_time", "impact_time", "post_time",
    "scene_profile", "extra_scene_tags",
    "notes",
]


def load_jsonl(path: Path) -> List[Dict]:
    rows = []
    with path.open("r", encoding="utf-8-sig") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert annotation jsonl to Excel workbook.")
    parser.add_argument("--input", required=True, help="Input jsonl")
    parser.add_argument("--output", required=True, help="Output xlsx")
    parser.add_argument("--sheet", default="annotation", help="Sheet name")
    args = parser.parse_args()

    rows = load_jsonl(Path(args.input).resolve())
    out = Path(args.output).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet

    lists = wb.create_sheet("lists")
    lists.sheet_state = "hidden"
    for i, v in enumerate(ACCIDENT_TYPES, start=2):
        lists[f"A{i}"] = v
    for i, v in enumerate(SCENE_TAGS, start=2):
        lists[f"B{i}"] = v
    for i, k in enumerate(SCENE_PROFILES.keys(), start=2):
        lists[f"C{i}"] = k

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.fill = header_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws["H1"].comment = Comment("先选场景模板（最省事）", "codex")
    ws["I1"].comment = Comment("模板不够时补充，逗号分隔，如：urban,crowded", "codex")

    for r in rows:
        raw_scene = r.get("scene_tags", []) or []
        if isinstance(raw_scene, list):
            scene = [str(x).strip() for x in raw_scene if str(x).strip()]
        else:
            scene = [x.strip() for x in str(raw_scene).split(",") if x.strip()]
        scene_set = set(scene)

        matched_profile = "custom"
        for profile, tags in SCENE_PROFILES.items():
            if profile == "custom":
                continue
            if set(tags).issubset(scene_set):
                matched_profile = profile
                break
        profile_tags = set(SCENE_PROFILES.get(matched_profile, []))
        extra_tags = [t for t in scene if t not in profile_tags]

        row = [
            r.get("sample_id", ""), r.get("split", ""), r.get("video", ""), r.get("accident_type", ""),
            r.get("onset_time", None), r.get("impact_time", None), r.get("post_time", None),
            matched_profile,
            ",".join(extra_tags),
            r.get("notes", ""),
        ]
        ws.append(row)

    max_row = ws.max_row
    if max_row >= 2:
        acc_dv = DataValidation(type="list", formula1="=lists!$A$2:$A$5", allow_blank=True)
        time_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="36000", allow_blank=True)
        scene_profile_dv = DataValidation(type="list", formula1="=lists!$C$2:$C$11", allow_blank=True)
        ws.add_data_validation(acc_dv)
        ws.add_data_validation(time_dv)
        ws.add_data_validation(scene_profile_dv)
        acc_dv.add(f"D2:D{max_row}")
        time_dv.add(f"E2:G{max_row}")
        scene_profile_dv.add(f"H2:H{max_row}")

    widths = {"A": 28, "B": 10, "C": 46, "D": 18, "E": 12, "F": 12, "G": 12, "H": 28, "I": 34, "J": 36}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J1" if max_row < 2 else f"A1:J{max_row}"
    wb.save(out)
    print(f"[DONE] rows={len(rows)} out={out}")


if __name__ == "__main__":
    main()
